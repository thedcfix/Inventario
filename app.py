import functools
import os

from io import BytesIO

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash

from config import Config

# ─── App factory ─────────────────────────────────────────────
app = Flask(__name__)
app.config.from_object(Config)

# ─── Azure services (lazy-initialised, local fallback) ──────
_cosmos = None
_storage = None


def _use_azure():
    """Return True if Azure credentials are configured."""
    return bool(app.config.get("COSMOS_ENDPOINT")) and bool(
        app.config.get("AZURE_STORAGE_CONNECTION_STRING")
    )


def get_cosmos():
    global _cosmos
    if _cosmos is None:
        if _use_azure():
            from services.cosmos_service import CosmosService

            _cosmos = CosmosService(
                endpoint=app.config["COSMOS_ENDPOINT"],
                key=app.config["COSMOS_KEY"],
                database_name=app.config["COSMOS_DATABASE"],
                container_name=app.config["COSMOS_CONTAINER"],
            )
        else:
            from services.local_service import LocalCosmosService

            _cosmos = LocalCosmosService()
    return _cosmos


def get_storage():
    global _storage
    if _storage is None:
        if _use_azure():
            from services.storage_service import StorageService

            _storage = StorageService(
                connection_string=app.config["AZURE_STORAGE_CONNECTION_STRING"],
                container_name=app.config["AZURE_STORAGE_CONTAINER"],
            )
        else:
            from services.local_service import LocalStorageService

            _storage = LocalStorageService()
    return _storage


# ─── Auth helpers ────────────────────────────────────────────
def login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def _refresh_photo_urls(items_or_item):
    """Regenerate SAS URLs for items stored in Azure (private container)."""
    if not _use_azure():
        return items_or_item
    storage = get_storage()

    def _fix_item(item):
        blob = item.get("blob_name", "")
        photo_url = item.get("photo_url", "")
        # Extract blob name from plain URL if blob_name field is missing
        if not blob and photo_url and "blob.core.windows.net" in photo_url:
            # URL format: https://account.blob.core.windows.net/container/blobname.jpg
            blob = photo_url.split("/")[-1].split("?")[0]
        if blob:
            item["photo_url"] = storage.get_photo_url(blob)
            if not item.get("blob_name"):
                item["blob_name"] = blob

    if isinstance(items_or_item, list):
        for item in items_or_item:
            _fix_item(item)
    elif isinstance(items_or_item, dict):
        _fix_item(items_or_item)
    return items_or_item


# ─── Auth routes ─────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if username == app.config["APP_USERNAME"] and check_password_hash(
            app.config["APP_PASSWORD_HASH"], password
        ):
            session["logged_in"] = True
            session["username"] = username
            return redirect(url_for("index"))

        return render_template("login.html", error="Credenziali non valide.")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── Dashboard ───────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    cosmos = get_cosmos()
    search = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    items = cosmos.get_items(search=search or None, category=category or None)
    _refresh_photo_urls(items)
    categories = cosmos.get_categories()
    total_price = sum(float(item.get("price", 0)) * int(item.get("quantity", 1)) for item in items)
    return render_template(
        "index.html",
        items=items,
        categories=categories,
        current_search=search,
        current_category=category,
        total_price=total_price,
    )


# ─── Export Excel ────────────────────────────────────────────
@app.route("/export")
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    cosmos = get_cosmos()
    search = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    items = cosmos.get_items(search=search or None, category=category or None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Nome", "Descrizione", "Categoria", "Quantità", "Prezzo unitario (€)", "Valore totale (€)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, item in enumerate(items, 2):
        price = float(item.get("price", 0))
        qty = int(item.get("quantity", 1))
        values = [
            item.get("name", ""),
            item.get("description", ""),
            item.get("category", ""),
            qty,
            price,
            price * qty,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col >= 5:
                cell.number_format = '#,##0.00'

    total_row = len(items) + 2
    ws.cell(row=total_row, column=4, value="TOTALE").font = Font(bold=True)
    # SUBTOTAL 109 = SUM that respects active filters
    ws.cell(row=total_row, column=6).value = f"=SUBTOTAL(109,F2:F{total_row - 1})"
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = [30, 40, 20, 12, 18, 18][col - 1]

    # Add Excel table with auto-filter
    last_data_row = len(items) + 1
    if last_data_row >= 2:
        table_ref = f"A1:F{last_data_row}"
        tab = Table(displayName="Inventario", ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "inventario"
    if category:
        filename += f"_{category}"
    if search:
        filename += f"_{search.replace(' ', '_')}"
    filename += ".xlsx"

    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── API: items JSON (for async search) ─────────────────────
@app.route("/api/items")
@login_required
def api_items():
    cosmos = get_cosmos()
    search = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    items = cosmos.get_items(search=search or None, category=category or None)
    return jsonify(items)


# ─── Create item ─────────────────────────────────────────────
@app.route("/item/new", methods=["GET", "POST"])
@login_required
def item_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        category = request.form.get("category", "").strip()
        new_category = request.form.get("new_category", "").strip()
        quantity = request.form.get("quantity", 1)
        price = request.form.get("price", 0)

        if new_category:
            category = new_category

        if not name or not category:
            cosmos = get_cosmos()
            return render_template(
                "item_form.html",
                mode="new",
                categories=cosmos.get_categories(),
                item=request.form,
                error="Nome e categoria sono obbligatori.",
            )

        photo_url = ""
        blob_name = ""
        photo = request.files.get("photo")
        if photo and photo.filename:
            storage = get_storage()
            blob_name, photo_url = storage.upload_photo(photo.stream, photo.filename)

        cosmos = get_cosmos()
        item = cosmos.create_item(
            name=name,
            description=description,
            category=category,
            quantity=quantity,
            photo_url=photo_url,
            price=price,
        )
        # Store blob_name for later deletion
        if blob_name:
            cosmos.update_item(item["id"], category, {"blob_name": blob_name})

        return redirect(url_for("index"))

    cosmos = get_cosmos()
    return render_template(
        "item_form.html", mode="new", categories=cosmos.get_categories(), item={}
    )


# ─── Item detail ─────────────────────────────────────────────
@app.route("/item/<item_id>")
@login_required
def item_detail(item_id):
    cosmos = get_cosmos()
    item = cosmos.get_item_by_id(item_id)
    if not item:
        abort(404)
    _refresh_photo_urls(item)
    return render_template("item_detail.html", item=item)


# ─── Edit item ───────────────────────────────────────────────
@app.route("/item/<item_id>/edit", methods=["GET", "POST"])
@login_required
def item_edit(item_id):
    cosmos = get_cosmos()
    item = cosmos.get_item_by_id(item_id)
    if not item:
        abort(404)
    _refresh_photo_urls(item)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        category = request.form.get("category", "").strip()
        new_category = request.form.get("new_category", "").strip()
        quantity = request.form.get("quantity", 1)
        price = request.form.get("price", 0)

        if new_category:
            category = new_category

        if not name or not category:
            return render_template(
                "item_form.html",
                mode="edit",
                categories=cosmos.get_categories(),
                item=item,
                error="Nome e categoria sono obbligatori.",
            )

        data = {
            "name": name,
            "description": description,
            "category": category,
            "quantity": int(quantity),
            "price": float(price),
        }

        # Handle photo update
        photo = request.files.get("photo")
        if photo and photo.filename:
            storage = get_storage()
            # Delete old photo
            old_blob = item.get("blob_name", "")
            if old_blob:
                storage.delete_photo(old_blob)
            blob_name, photo_url = storage.upload_photo(photo.stream, photo.filename)
            data["photo_url"] = photo_url
            data["blob_name"] = blob_name

        cosmos.update_item(item_id, category, data)

        return redirect(url_for("item_detail", item_id=item_id))

    return render_template(
        "item_form.html", mode="edit", categories=cosmos.get_categories(), item=item
    )


# ─── Delete item ─────────────────────────────────────────────
@app.route("/item/<item_id>/delete", methods=["POST"])
@login_required
def item_delete(item_id):
    cosmos = get_cosmos()
    item = cosmos.get_item_by_id(item_id)
    if not item:
        abort(404)

    # Delete photo from blob storage
    blob_name = item.get("blob_name", "")
    photo_url = item.get("photo_url", "")
    # Extract blob name from URL if field is missing
    if not blob_name and photo_url:
        if "blob.core.windows.net" in photo_url:
            blob_name = photo_url.split("/")[-1].split("?")[0]
        elif "/static/uploads/" in photo_url:
            blob_name = photo_url.split("/")[-1]
    if blob_name:
        storage = get_storage()
        storage.delete_photo(blob_name)

    cosmos.delete_item(item_id, item["category"])
    return redirect(url_for("index"))


# ─── Error handlers ──────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return render_template("base.html", error_code=404, error_msg="Pagina non trovata"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("base.html", error_code=500, error_msg="Errore del server"), 500


# ─── CLI helper: generate password hash ─────────────────────
@app.cli.command("generate-hash")
def generate_hash():
    """Generate a password hash for APP_PASSWORD_HASH."""
    import getpass

    password = getpass.getpass("Password: ")
    print(generate_password_hash(password))


if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1", host="0.0.0.0", port=5000)
